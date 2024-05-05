from selenium.webdriver.common.by import By
from homepagechange import Homepagechange
from searchpagechange import Searchpagechange
from formpagechange import Formpagechange
from affectedcipagechange import Affectedcipagechange
import logging.config
import logging
import pytest
import time
import openpyxl

@pytest.mark.usefixtures("test_initial_setup")
class TestCopyChange:

    logging.config.fileConfig('pytest.ini')
    my_logger = logging.getLogger()
    dir_path = 'C:/Users/VNarasi7/Servicenowautomate/data/Sample_Input_Latest_New.xlsx'

    def test_launchURL(self):
        
        hpc = Homepagechange(self.driver)
        logging.info("The object is created for Homepage")
        logging.info("The homepage function is called")
        hpc.homepage_func()
        time.sleep(5)
        parent_window = self.driver.current_window_handle
        logging.info("The title of the current page is: "+self.driver.title)
        self.driver.switch_to.window(self.driver.window_handles[1])
        logging.info("The title of the current page is: "+self.driver.title)
        time.sleep(2)
        try:
            workbook = openpyxl.load_workbook(self.dir_path)
            logging.info("The workbook is loaded")
            CRSheet = workbook['ChangeRequest']
            logging.info("The ChangeRquest worksheet is loaded")
            total_rows = CRSheet.max_row
            total_column = CRSheet.max_column
            logging.info("CR total_rows: %s", total_rows)
            logging.info("CR total_columns: %s", total_column)
            ACISheet = workbook['AffectedCI']
            logging.info("The AffectedCI worksheet is loaded")
            ACItotal_rows = ACISheet.max_row
            ACItotal_column = ACISheet.max_column
            logging.info("AffectedCI total_rows: %s", ACItotal_rows)
            logging.info("AffectedCI total_columns: %s", ACItotal_column)
        except FileNotFoundError:
            logging.info("The excel file is not found")

        spc = Searchpagechange(self.driver)
        logging.info("The object is created for Search page")
        r = 2
        while(r<total_rows+1):
            c = 1
            searchReqValue = CRSheet.cell(r,c).value
            urgencyValue = CRSheet.cell(r,c+1).value
            shortdescValue = CRSheet.cell(r,c+2).value
            descValue = CRSheet.cell(r,c+3).value
            assignciValue = CRSheet.cell(r,c+4).value
            assigngrpValue = CRSheet.cell(r,c+5).value
            justifyValue = CRSheet.cell(r,c+6).value
            logging.info("The search page function is called")
            if(spc.Searchpage_func(searchReqValue)):

                fpc = Formpagechange(self.driver)
                logging.info("The object is created for form page")
                logging.info("The form page function is called")                
                newChgRequestNumber = fpc.formpage_func(urgencyValue, shortdescValue, descValue, assignciValue, assigngrpValue, justifyValue)
                if(newChgRequestNumber!=None):

                    CRSheet.cell(r,c+7).value = newChgRequestNumber
                    workbook.save(self.dir_path)
                    logging.info("The new change request is saved in sheet of the workbook")
                    afci = Affectedcipagechange(self.driver)
                    logging.info("The object is created for affectedci page ")
                    logging.info("The affectedci function is called ")
                    afci.affectedci_func(ACISheet, r) 
                    
                else:
                    CRSheet.cell(r,c+7).value = "Failed" 
                    workbook.save(self.dir_path)
                    logging.error("Failed! The new change request is not created for the following change")      
                self.driver.switch_to.default_content()
            else:
                CRSheet.cell(r,c+7).value = "Failed" 
                workbook.save(self.dir_path)
                logging.error("Failed! Unable to search for the following change request. Trying for the next change request")
            r=r+1
